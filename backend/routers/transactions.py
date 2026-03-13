from datetime import date
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy import select, func, text
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from backend.database import get_db
from backend.models import Account, Transaction, Category
from backend.schemas.transaction import TransactionOut, TransactionUpdate

router = APIRouter(prefix="/api/transactions", tags=["transactions"])


@router.get("", response_model=list[TransactionOut])
async def list_transactions(
    account_id: int | None = None,
    category_id: int | None = None,
    uncategorized: bool = False,
    search: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    sort: str = Query(default="date", pattern="^(date|description|amount|balance_after|category)$"),
    sort_dir: str = Query(default="desc", pattern="^(asc|desc)$"),
    limit: int = Query(default=100, le=500),
    offset: int = 0,
    db: AsyncSession = Depends(get_db),
):
    stmt = select(Transaction).options(joinedload(Transaction.category))

    if account_id is not None:
        stmt = stmt.where(Transaction.account_id == account_id)
    if uncategorized:
        stmt = stmt.where(Transaction.category_id.is_(None))
    elif category_id is not None:
        stmt = stmt.where(Transaction.category_id == category_id)
    if search:
        stmt = stmt.where(Transaction.description.ilike(f"%{search}%"))
    if start_date is not None:
        stmt = stmt.where(Transaction.date >= start_date)
    if end_date is not None:
        stmt = stmt.where(Transaction.date <= end_date)

    sort_columns = {
        "date": [Transaction.date, Transaction.id],
        "description": [Transaction.description],
        "amount": [Transaction.amount],
        "balance_after": [Transaction.balance_after],
        "category": [Category.name],
    }
    cols = sort_columns.get(sort, [Transaction.date, Transaction.id])
    if sort_dir == "desc":
        stmt = stmt.order_by(*[c.desc() for c in cols])
    else:
        stmt = stmt.order_by(*[c.asc() for c in cols])
    stmt = stmt.offset(offset).limit(limit)

    result = await db.execute(stmt)
    txns = result.scalars().all()

    return [
        TransactionOut(
            id=t.id,
            account_id=t.account_id,
            date=t.date,
            description=t.description,
            amount=t.amount,
            currency=t.currency,
            amount_hkd=t.amount_hkd,
            balance_after=t.balance_after,
            category_id=t.category_id,
            category_source=t.category_source,
            category_confidence=t.category_confidence,
            category_name=t.category.name if t.category else None,
            transfer_pair_id=t.transfer_pair_id,
        )
        for t in txns
    ]


class BulkCategoryUpdate(BaseModel):
    transaction_ids: list[int]
    category_id: int


class BulkUpdateResponse(BaseModel):
    updated: int


@router.post("/bulk-categorize", response_model=BulkUpdateResponse)
async def bulk_categorize(data: BulkCategoryUpdate, db: AsyncSession = Depends(get_db)):
    """Assign a category to multiple transactions at once."""
    if not data.transaction_ids:
        return BulkUpdateResponse(updated=0)

    result = await db.execute(
        select(Transaction).where(Transaction.id.in_(data.transaction_ids))
    )
    txns = result.scalars().all()

    for txn in txns:
        txn.category_id = data.category_id
        txn.category_source = "manual"
        txn.category_confidence = 1.0

    await db.commit()
    return BulkUpdateResponse(updated=len(txns))


@router.get("/export")
async def export_transactions(
    account_id: int | None = None,
    category_id: int | None = None,
    uncategorized: bool = False,
    search: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    exclude_transfers: bool = False,
    sort: str = Query(default="date", pattern="^(date|description|amount|balance_after|category)$"),
    sort_dir: str = Query(default="desc", pattern="^(asc|desc)$"),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(Transaction).options(joinedload(Transaction.category))

    if account_id is not None:
        stmt = stmt.where(Transaction.account_id == account_id)
    if uncategorized:
        stmt = stmt.where(Transaction.category_id.is_(None))
    elif category_id is not None:
        stmt = stmt.where(Transaction.category_id == category_id)
    if search:
        stmt = stmt.where(Transaction.description.ilike(f"%{search}%"))
    if start_date is not None:
        stmt = stmt.where(Transaction.date >= start_date)
    if end_date is not None:
        stmt = stmt.where(Transaction.date <= end_date)
    if exclude_transfers:
        stmt = stmt.where(Transaction.transfer_pair_id.is_(None))

    sort_columns = {
        "date": [Transaction.date, Transaction.id],
        "description": [Transaction.description],
        "amount": [Transaction.amount],
        "balance_after": [Transaction.balance_after],
        "category": [Category.name],
    }
    cols = sort_columns.get(sort, [Transaction.date, Transaction.id])
    if sort_dir == "desc":
        stmt = stmt.order_by(*[c.desc() for c in cols])
    else:
        stmt = stmt.order_by(*[c.asc() for c in cols])

    # Load account names
    acct_result = await db.execute(select(Account))
    acct_map = {a.id: a.name for a in acct_result.scalars().all()}

    result = await db.execute(stmt)
    txns = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["Date", "Account", "Description", "Currency", "Amount", "Balance After", "Category"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    for t in txns:
        ws.append([
            t.date.isoformat(),
            acct_map.get(t.account_id, ""),
            t.description,
            t.currency,
            float(t.amount),
            float(t.balance_after) if t.balance_after is not None else None,
            t.category.name if t.category else "",
        ])

    # Auto-width columns
    for col_idx, _ in enumerate(headers, 1):
        letter = get_column_letter(col_idx)
        max_len = max(len(str(cell.value or "")) for cell in ws[letter])
        ws.column_dimensions[letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "transactions"
    if start_date:
        filename += f"_{start_date}"
    if end_date:
        filename += f"_to_{end_date}"
    filename += ".xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.patch("/{transaction_id}", response_model=TransactionOut)
async def update_transaction(
    transaction_id: int,
    update: TransactionUpdate,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Transaction)
        .options(joinedload(Transaction.category))
        .where(Transaction.id == transaction_id)
    )
    txn = result.scalar_one_or_none()
    if txn is None:
        raise HTTPException(404, "Transaction not found")

    if update.category_id is not None:
        txn.category_id = update.category_id
        txn.category_source = update.category_source
        txn.category_confidence = 1.0 if update.category_source == "manual" else None

    if update.is_transfer is not None:
        txn.transfer_pair_id = -1 if update.is_transfer else None

    await db.commit()
    await db.refresh(txn, ["category"])

    return TransactionOut(
        id=txn.id,
        account_id=txn.account_id,
        date=txn.date,
        description=txn.description,
        amount=txn.amount,
        currency=txn.currency,
        amount_hkd=txn.amount_hkd,
        balance_after=txn.balance_after,
        category_id=txn.category_id,
        category_source=txn.category_source,
        category_confidence=txn.category_confidence,
        category_name=txn.category.name if txn.category else None,
        transfer_pair_id=txn.transfer_pair_id,
    )


class DetectTransfersResponse(BaseModel):
    linked: int
    total_pairs: int


class DeduplicateResponse(BaseModel):
    deleted: int


@router.post("/deduplicate", response_model=DeduplicateResponse)
async def deduplicate_transactions(db: AsyncSession = Depends(get_db)):
    """Delete duplicate transactions, keeping the one with the lowest id."""
    result = await db.execute(text("""
        SELECT id FROM transactions
        WHERE id NOT IN (
            SELECT MIN(id)
            FROM transactions
            GROUP BY account_id, date, description, amount, currency
        )
    """))
    dup_ids = [row[0] for row in result.all()]

    if dup_ids:
        # Clear transfer_pair_id references to duplicates first
        id_list = ",".join(str(i) for i in dup_ids)
        await db.execute(text(
            f"UPDATE transactions SET transfer_pair_id = NULL WHERE transfer_pair_id IN ({id_list})"
        ))
        await db.execute(text(
            f"DELETE FROM transactions WHERE id IN ({id_list})"
        ))
        await db.commit()

    return DeduplicateResponse(deleted=len(dup_ids))


@router.post("/detect-transfers", response_model=DetectTransfersResponse)
async def detect_transfers(db: AsyncSession = Depends(get_db)):
    """Find and link internal transfers."""
    # Clear existing transfer links
    await db.execute(
        text("UPDATE transactions SET transfer_pair_id = NULL WHERE transfer_pair_id IS NOT NULL")
    )

    # Get credit card account IDs
    cc_result = await db.execute(
        select(Account.id).where(Account.account_type == "credit_card")
    )
    cc_ids = {row[0] for row in cc_result.all()}

    # Step 1: Exact match pairs (same amount, opposite sign, different account, <=1 day)
    result = await db.execute(text("""
        SELECT t1.id, t2.id
        FROM transactions t1
        JOIN transactions t2 ON t1.id < t2.id
            AND t1.account_id != t2.account_id
            AND abs(t1.amount + t2.amount) < 0.01
            AND t1.amount != 0
            AND abs(julianday(t1.date) - julianday(t2.date)) <= 1
            AND t1.currency = t2.currency
        ORDER BY t1.date DESC
    """))
    pairs = result.all()

    used: set[int] = set()
    linked = 0
    for id1, id2 in pairs:
        if id1 in used or id2 in used:
            continue
        await db.execute(
            text("UPDATE transactions SET transfer_pair_id = :pair WHERE id = :id"),
            {"id": id1, "pair": id2},
        )
        await db.execute(
            text("UPDATE transactions SET transfer_pair_id = :pair WHERE id = :id"),
            {"id": id2, "pair": id1},
        )
        used.add(id1)
        used.add(id2)
        linked += 1

    # Step 2: Credit card — any positive amount (incoming) is a repayment/transfer
    # Mark unlinked ones with transfer_pair_id = -1 (self-referencing sentinel)
    if cc_ids:
        cc_list = ",".join(str(i) for i in cc_ids)
        result = await db.execute(text(f"""
            SELECT id FROM transactions
            WHERE account_id IN ({cc_list})
            AND amount > 0
            AND transfer_pair_id IS NULL
        """))
        unlinked_cc = result.all()
        for (tid,) in unlinked_cc:
            await db.execute(
                text("UPDATE transactions SET transfer_pair_id = -1 WHERE id = :id"),
                {"id": tid},
            )
            linked += 1

    await db.commit()
    return DetectTransfersResponse(linked=linked, total_pairs=linked)
